"""도메인별 성분 원천 서류 → 정규화 폼(asap-coi-v1 계열) 파서 레지스트리.

도메인별 서류 → 채우는 것 → 직결 분기 (설계표 기준):
  식품(01~24)      COI/성분명세서   entries(표기순)·%·알레르겐   ✅ 구현(coi_normalize)
  화학(28~38)      COA(시험성적서)  순도 assay %·성분·CAS 번호   ⬜ 스텁
  금속(72~83)      MTC(mill test)  합금 성분 %(C·Cr·Ni…)        ⬜ 스텁
  섬유(50~63)      혼용률 표        섬유 성분 %                  ⬜ 스텁
  전기·기계(84~87) 스펙시트         정격 V·A·W·cc·용량           ⬜ 스텁
  화장품(33)       전성분(INCI)     성분 리스트                  ⬜ 스텁
  플라스틱(39)     수지 명세        폴리머 종류                  ⬜ 스텁

폼 계약(공통 코어): entries[{name_ko, name_en, order_index, percent,
role, section}] + 도메인 확장 필드(화학 cas_no, 금속 alloy_grade,
섬유 fiber_type, 전기 rated_*). 스텁은 빈 폼(parse_status='stub')을
반환한다 — 경로·계약을 먼저 고정하고 구현은 도메인별 샘플 서류 확보 후.

주입 시점 메모: 현재 주입은 PU 직후(도메인 라우팅 전)다. 챕터별 상이한
서류를 정밀하게 받으려면 DomainRouter 뒤가 자연스럽다는 지적이 있는데,
서류 '유형' 판별은 서류 자체의 헤더 시그니처로도 가능해서(아래 _sniff)
순서 재배치 없이 굴러가는 절충이 있다 — 결정은 보류(설계자 고민 항목).
"""
from __future__ import annotations

from pathlib import Path
from typing import Callable


def _empty_form(domain: str, path: Path, status: str = "stub") -> dict:
    return {
        "coi_form_version": "asap-coi-v1",
        "domain": domain,
        "source_file": str(path),
        "product_name_hint": path.stem,
        "sections": [],
        "entries": [],
        "principal_candidates": [],
        "principal_ingredient": "",
        "accessory_ingredients": [],
        "coverage": "empty",
        "parse_status": status,
    }


def ParseFoodCoi(path: Path) -> dict:
    """식품 COI — 구현체는 coi_normalize.NormalizeCoiFile (위임)."""
    from coi_normalize import NormalizeCoiFile  # DB/sources 동일 디렉토리
    form = NormalizeCoiFile(path)
    form["domain"] = "food"
    return form


def _read_cells(path: Path) -> list[list[str]]:
    """xlsx/txt → 행 단위 셀 텍스트 (파서 공용 원료)."""
    if path.suffix.lower() in (".txt", ".csv"):
        try:
            return [[line] for line in path.read_text(encoding="utf-8").splitlines()]
        except Exception:  # noqa: BLE001
            return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        rows = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
        return rows
    except Exception:  # noqa: BLE001
        return []


_CAS_RE = __import__("re").compile(r"\b(\d{2,7}-\d{2}-\d)\b")
_PCT_RE = __import__("re").compile(r"(\d{1,3}(?:[.,]\d{1,3})?)\s*%")


def ParseChemicalCoa(path: Path) -> dict:
    """화학 COA(시험성적서): CAS 번호가 있는 행을 entries로 — 순도 assay %
    동반 캡처. CAS는 ECICS 대조 키(29류 분기의 사실상 전부)."""
    form = _empty_form("chemical", path, status="ok")
    order = 0
    for row in _read_cells(path):
        joined = " | ".join(c for c in row if c)
        cas = _CAS_RE.search(joined)
        if not cas:
            continue
        pct = _PCT_RE.search(joined)
        name = next((c.strip() for c in row
                     if c.strip() and not _CAS_RE.fullmatch(c.strip())
                     and len(c.strip()) >= 3), "")
        order += 1
        form["entries"].append({
            "order_index": order, "section": "assay", "role": "other",
            "name_ko": name, "name_raw": joined[:120], "name_en": name,
            "sub_ingredients": [], "origin": "",
            "percent": float(pct.group(1).replace(",", ".")) if pct else None,
            "cas_no": cas.group(1),
        })
    form["parse_status"] = "ok" if form["entries"] else "empty"
    form["coverage"] = "full" if form["entries"] else "empty"
    return form


def ParseMetalMtc(path: Path) -> dict:
    """금속 MTC(mill test certificate): 합금 성분 %(C·Cr·Ni…) → 72류
    스테인리스/합금강 분기. 확장 필드: alloy_grade, entries[]=원소별 %."""
    return _empty_form("metal", path)


def ParseTextileFiberContent(path: Path) -> dict:
    """섬유 혼용률 표: 섬유 성분 % → 커튼 6303.91/92(면/합성),
    바지 6203.49 등 US_KR 오답 축. 확장 필드: fiber_type."""
    return _empty_form("textile", path)


def ParseElectronicsSpecSheet(path: Path) -> dict:
    """전기·기계·차량 스펙시트: 정격 V·A·W·cc·용량 — quant형 전멸 구간의
    답. 확장 필드: rated_voltage/current/power/displacement/capacity."""
    return _empty_form("electronics", path)


def ParseCosmeticsInci(path: Path) -> dict:
    """화장품 전성분(INCI): 콤마 연속 나열 라인을 성분 리스트로.

    INCI는 국제명명(영문)이라 번역 불필요 — name_en=자기 자신. 표기순이
    법정 함량 내림차순(1% 초과 구간)이라 order_index가 곧 서열 사실.
    규제 접점: entries가 Annex II 금지물질 스크리닝(DB의
    ingredient_prohibited_check 요건)의 입력이 된다."""
    form = _empty_form("cosmetics", path, status="ok")
    best_line, best_n = "", 0
    for row in _read_cells(path):
        for cell in row:
            n = cell.count(",")
            ascii_ratio = sum(ch.isascii() for ch in cell) / max(1, len(cell))
            if n >= 4 and ascii_ratio > 0.7 and n > best_n:
                best_line, best_n = cell, n
    import re as _re
    best_line = _re.sub(r"^\s*(전성분|성분|ingredients?)\s*[:：]\s*", "", best_line, flags=_re.I)
    for i, name in enumerate(s.strip() for s in best_line.split(",")):
        if len(name) < 2:
            continue
        form["entries"].append({
            "order_index": i + 1, "section": "inci", "role": "other",
            "name_ko": name, "name_raw": name, "name_en": name,
            "sub_ingredients": [], "percent": None, "origin": "",
        })
    form["parse_status"] = "ok" if form["entries"] else "empty"
    form["coverage"] = "full" if form["entries"] else "empty"
    return form


def ParsePlasticsResinSpec(path: Path) -> dict:
    """플라스틱 수지 명세: 폴리머 종류 → 39류 폴리머별 분기."""
    return _empty_form("plastics", path)


# 도메인 → (담당 챕터, 파서). 챕터는 안내용 메타 — 주입 자체는 폼이
# 성립하면 도메인 무관하게 동일 계약(composition_lane)으로 흐른다.
DOMAIN_PARSERS: dict[str, tuple[str, Callable[[Path], dict]]] = {
    "food": ("01-24", ParseFoodCoi),
    "chemical": ("28-38", ParseChemicalCoa),
    "metal": ("72-83", ParseMetalMtc),
    "textile": ("50-63", ParseTextileFiberContent),
    "electronics": ("84-87", ParseElectronicsSpecSheet),
    "cosmetics": ("33", ParseCosmeticsInci),
    "plastics": ("39", ParsePlasticsResinSpec),
}


def _sniff(path: Path) -> str:
    """서류 자체의 시그니처로 도메인 추정 — DomainRouter 없이도 서류
    유형을 가릴 수 있게 하는 절충 지점. 스텁: 파일명 키워드만."""
    name = path.name.lower()
    if any(k in name for k in ("coa", "시험성적", "성적서")):
        return "chemical"
    if any(k in name for k in ("mtc", "mill", "성분증명")):
        return "metal"
    if any(k in name for k in ("혼용", "fiber", "섬유")):
        return "textile"
    if any(k in name for k in ("spec", "datasheet", "스펙")):
        return "electronics"
    if any(k in name for k in ("inci", "전성분")):
        return "cosmetics"
    if any(k in name for k in ("수지", "resin", "폴리머")):
        return "plastics"
    return "food"


def ParseDocument(path: Path, domain: str | None = None) -> dict:
    """단일 진입점: 도메인 명시 없으면 서류 시그니처로 추정."""
    resolved = domain or _sniff(path)
    _chapters, parser = DOMAIN_PARSERS.get(resolved, DOMAIN_PARSERS["food"])
    return parser(path)
