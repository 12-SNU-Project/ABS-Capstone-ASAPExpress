"""COI 로더 — 정규화 COI 폼(asap-coi-v1)의 로딩·주입이 주 기능.

[신형 — 주 경로] 서비스 전제는 "COI를 입력 서류로 받는다":
  업로드 COI → 정규화·번역 → coi_form/coi_product_map
  → 본 모듈이 PostgreSQL에서 찾아 composition_lane에 가산 주입.
  게이트: ASAP_COI_FORM_DIR=0. 주입 규칙(교차 합의제·기존값 보존·
  sauce_only 주성분 금지)은 coi50 스모크에서 실측 검증.

[파싱 유틸] ParseCoiComposition/_FindTableSheet 등은 정규화기의 원료.

[구식 — deprecated] LoadCoiEvidence(ASAP_COI_ROOT 디렉토리 스캔)는
  "COI를 미리 보유"하던 시절의 경로. 기존 수출품 개정 재검토 때만
  쓰일 수 있어 게이트 잠김 상태로만 유지한다.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable


import os
from bussiness_logic.core.runtime_asset_repository import (
    LoadCoiForms,
    LoadCoiProductMap,
)

PROJECT_ROOT = Path(__file__).resolve().parents[4]
# ASAP_COI_ROOT로 오버라이드 가능 — 실물 46파일은 현재 ~/ASAP_A/test에 있다
DEFAULT_COI_ROOT = Path(
    os.environ.get("ASAP_COI_ROOT")
    or PROJECT_ROOT / "test" / "COI(식품원재료풀이)"
)
TOKEN_RE = re.compile(r"[0-9A-Za-z가-힣]+")
INDEX_PREFIX_RE = re.compile(r"^\s*([0-9,\s]+)\.")


@dataclass(frozen=True, slots=True)
class CoiEvidence:
    path: Path
    text: str
    matchedScore: int = 0


def NormalizeText(value: object) -> str:
    return unicodedata.normalize("NFC", " ".join(str(value or "").split()))


def EvidenceTokens(text: str) -> set[str]:
    return {
        token.lower()
        for token in TOKEN_RE.findall(NormalizeText(text))
        if len(token) >= 2
    }


def _IndicesFromParent(path: Path) -> set[int]:
    match = INDEX_PREFIX_RE.match(path.parent.name)
    if match is None:
        return set()
    return {
        int(chunk)
        for chunk in match.group(1).replace(" ", "").split(",")
        if chunk.isdigit()
    }


def _XlsxPaths(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(
        path
        for path in root.rglob("*.xlsx")
        if not path.name.startswith("~$")
    )


def _ScorePath(path: Path, *, caseIndex: int | None, productName: str) -> int:
    score = 0
    if caseIndex is not None and caseIndex in _IndicesFromParent(path):
        score += 100
    haystack = f"{path.parent.name} {path.stem}"
    score += len(EvidenceTokens(productName) & EvidenceTokens(haystack)) * 10
    if "수출가능" in path.name:
        score += 4
    if "수정" in path.name or "최종" in path.name:
        score += 2
    if "복사" in path.name or "copy" in path.name.lower():
        score -= 1
    return score


def FindCoiPath(
    *,
    caseIndex: int | None,
    productName: str,
    coiRoot: Path = DEFAULT_COI_ROOT,
) -> Path | None:
    scored = [
        (_ScorePath(path, caseIndex=caseIndex, productName=productName), path)
        for path in _XlsxPaths(coiRoot)
    ]
    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], str(item[1])))
    score, path = scored[0]
    return path if score > 0 else None


def _CellToText(value: object) -> str:
    if value is None:
        return ""
    text = NormalizeText(value)
    if not text or text.lower() in {"none", "nan"}:
        return ""
    return text


def FlattenXlsxText(
    path: Path,
    *,
    maxChars: int = 8000,
    maxRowsPerSheet: int = 220,
) -> str:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to read COI xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet in workbook.worksheets:
            rowsAdded = 0
            parts.append(f"[sheet] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [_CellToText(value) for value in row]
                values = [value for value in values if value]
                if not values:
                    continue
                parts.append(" | ".join(values))
                rowsAdded += 1
                if rowsAdded >= maxRowsPerSheet:
                    parts.append("[sheet_truncated]")
                    break
                if sum(len(part) + 1 for part in parts) >= maxChars:
                    return "\n".join(parts)[:maxChars]
    finally:
        workbook.close()
    return "\n".join(parts)[:maxChars]


@lru_cache(maxsize=256)
def _CachedFlatten(pathText: str, maxChars: int) -> str:
    return FlattenXlsxText(Path(pathText), maxChars=maxChars)


# [11회차-2 §1-A] LoadCoiEvidence 소멸 — 구식 디렉토리 스캔 경로(상품명
# 유사도로 xlsx를 고르던 방식)는 정규화 폼 단일화(ParseCoiComposition ·
# InjectIntoProductUnderstanding)로 대체. 삭제 근거: 최근 2런 58건에서
# 구식 경로 서명 0 · 외부 소비자 0(product_understanding 호출부 동시
# 제거). FindCoiPath/DEFAULT_COI_ROOT는 진단 유틸 SummarizeCoiMatches가
# 계속 쓰므로 존치.


_HEADER_KEYS = ("품명", "성분", "함량")
_PERCENT_RE = re.compile(r"(\d+(?:[.,]\d+)?)\s*%?")


def _FindTableSheet(workbook, sheetName: str | None = None):
    """유효 헤더(품명·성분·함량)를 가진 시트 중 데이터 행이 가장 많은 것.

    실물 COI는 다중 시트에 예시 시트가 섞여 있다(1~2번째가 작성 예시인
    파일 실측). 시트명에 '예시'가 있으면 제외하고, 헤더 시그니처로 표를
    찾은 뒤 데이터 행수가 최대인 시트를 고른다 — 순수 결정론 선별.
    """
    best = None  # (data_rows, sheet, header_row_idx, colmap)
    for sheet in workbook.worksheets:
        if "예시" in str(sheet.title):
            continue
        # [07-20] 제품별 시트 분리 — 한 파일에 상품별 시트(설기·콩찰떡)가
        # 공존하는 실물. 시트 지정 시 그 시트만 파싱(정규화기 분리 산출용).
        if sheetName and str(sheet.title) != sheetName:
            continue
        header_idx, colmap = None, {}
        fb_idx, fb_map = None, {}
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            cells = {j: NormalizeText(v).lower() for j, v in enumerate(row) if v is not None}
            hits: dict[str, int] = {}
            for j, v in cells.items():
                # 실물 변형 흡수: 영문 헤더(목란), N차 원재료 열(차려낸)
                if ("품명" in v or "product name" in v or v.startswith("구분")) and "품명" not in hits:
                    hits["품명"] = j
                if ("성분" in v or v == "ingredient" or "1차 원재료" in v) and "성분" not in hits:
                    hits["성분"] = j
                if ("함량" in v or "content" in v) and "함량" not in hits:
                    hits["함량"] = j
            if "품명" in hits and "성분" in hits:
                header_idx = i
                # 재료명(하위 단계)·원산지 열도 있으면 기록
                colmap = hits
                for j, v in cells.items():
                    if "재료명" in v or "breakdown" in v:
                        colmap.setdefault("재료명", j)
                    if "원산지" in v or "origin" in v:
                        colmap.setdefault("원산지", j)
                break
            # [07-20] 폴백 헤더 — 품명 열 없는 2단 작성표 실물(전복미역국:
            # r7 '성분/영문/함량/1차 원료명', 품명은 별도 행). 성분∧함량만
            # 있는 첫 행을 예비 헤더로 기억, 정헤더 부재 시에만 사용.
            if (fb_idx is None and "성분" in hits and "함량" in hits):
                fb_idx, fb_map = i, dict(hits)
                for j, v in cells.items():
                    if "재료명" in v or "breakdown" in v:
                        fb_map.setdefault("재료명", j)
                    if "원산지" in v or "origin" in v:
                        fb_map.setdefault("원산지", j)
            if i > 120:
                break
        if header_idx is None and fb_idx is not None:
            header_idx, colmap = fb_idx, fb_map
        if header_idx is None:
            continue
        data_rows = 0
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i <= header_idx:
                continue
            if any(v is not None for v in row):
                data_rows += 1
            if i > header_idx + 400:
                break
        if best is None or data_rows > best[0]:
            best = (data_rows, sheet, header_idx, colmap)
    return best


def ParseCoiComposition(path: Path, *, maxEntries: int = 60,
                        sheetName: str | None = None) -> list[dict]:
    """COI 원료풀이 표 → 구조화 성분 엔트리 (표기순 = 함량 내림차순 규칙).

    반환: [{ingredient_name, component, percent, order_index, origin,
            source: "coi"}] — composition lane의 ingredient_entries와
    같은 소비 계약. 실패는 빈 목록(no-op).
    """
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:  # noqa: BLE001 — 손상 파일 = 증거 없음
        return []
    try:
        found = _FindTableSheet(workbook, sheetName=sheetName)
        if not found:
            return []
        _n, sheet, header_idx, colmap = found
        c_comp = colmap.get("품명")
        c_ing = colmap.get("성분")
        c_pct = colmap.get("함량")
        c_org = colmap.get("원산지")
        entries: list[dict] = []
        component = ""
        order = 0
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i <= header_idx or len(entries) >= maxEntries:
                continue
            def cell(j):
                if j is None or j >= len(row):
                    return ""
                return NormalizeText(row[j])
            comp_v, ing_v = cell(c_comp), cell(c_ing)
            if comp_v:
                component = comp_v.replace("\n", " ")
            if not ing_v:
                continue  # 하위 단계(재료명)만 있는 행은 1차 성분이 아님
            percent = None
            m = _PERCENT_RE.search(cell(c_pct))
            if m:
                try:
                    percent = float(m.group(1).replace(",", "."))
                except ValueError:
                    percent = None
            order += 1
            entries.append({
                "ingredient_name": ing_v,
                "component": component,
                "percent": percent,
                "order_index": order,
                "origin": cell(c_org),
                "source": "coi",
            })
        # [07-20] 분율 함량 정규화 — 작성표가 %를 소수 분율로 담는 실물
        # (전복미역국 자숙전복 0.04 = 4%). 전량이 1 이하일 때만 ×100 —
        # 기존 백분율 폼(33.3 등)은 무접촉(이산 판정).
        pcts = [e["percent"] for e in entries
                if isinstance(e.get("percent"), (int, float))]
        if pcts and max(pcts) <= 1.0:
            for e in entries:
                if isinstance(e.get("percent"), (int, float)):
                    e["percent"] = round(e["percent"] * 100, 2)
        return entries
    finally:
        workbook.close()


def SummarizeCoiMatches(
    rows: Iterable[tuple[int | None, str]],
    *,
    coiRoot: Path = DEFAULT_COI_ROOT,
) -> list[dict[str, object]]:
    return [
        {
            "case_index": caseIndex,
            "product_name": productName,
            "coi_path": str(
                FindCoiPath(
                    caseIndex=caseIndex,
                    productName=productName,
                    coiRoot=coiRoot,
                )
                or ""
            ),
        }
        for caseIndex, productName in rows
    ]


# ══════════════════════════════
# 신형: 정규화 COI 폼 로딩·주입 (주 경로)
# ══════════════════════════════
from typing import Any

def _forms_enabled() -> bool:
    """Normalized COI is default-on and database-backed."""
    raw = (os.environ.get("ASAP_COI_FORM_DIR") or "").strip().lower()
    return raw not in ("0", "off", "false")


def _nfc(t: object) -> str:
    return unicodedata.normalize("NFC", str(t or "")).replace("\xa0", " ").strip()


def _merge_forms(forms: list[dict]) -> dict:
    merged = {
        "sections": [], "entries": [],
        "principal_candidates": [], "principal_candidates_en": [],
        "accessory_ingredients": [],
        "coverage": "full" if any(f.get("coverage") == "full" for f in forms) else "sauce_only",
        # [07-20 Task A] 구성품 병합 표식 — 밀키트(어묵+김치+면 등 별
        # 제조사 서류 결합)의 % 는 **구성품 내부 비율**(어묵 90%는 어묵
        # 봉지 안에서 90%지 제품의 90% 아님)이라 제품 레벨 quant로 쓰면
        # 오독한다. 구성품 상호 비율 데이터가 없으므로 quant 미주입(미결
        # > 오판 — 원리 8). 단일 폼(len==1)은 표식 없음(정상 % 주입).
        "_constituent_merge": len(forms) > 1,
    }
    order = 0
    for f in forms:
        hint = str(f.get("product_name_hint") or "")[:14]
        for e in f.get("entries") or []:
            order += 1
            e2 = dict(e)
            e2["order_index"] = order
            e2["section"] = f"{hint}:{e.get('section') or ''}".strip(":")
            merged["entries"].append(e2)
        if f.get("coverage") == "full":
            cands = f.get("principal_candidates") or []
            cands_en = f.get("principal_candidates_en") or []
            for i, c in enumerate(cands):
                if c not in merged["principal_candidates"]:
                    merged["principal_candidates"].append(c)
                    merged["principal_candidates_en"].append(
                        cands_en[i] if i < len(cands_en) else "")
        for a in f.get("accessory_ingredients") or []:
            if a not in merged["accessory_ingredients"]:
                merged["accessory_ingredients"].append(a)
    merged["principal_candidates"] = merged["principal_candidates"][:4]
    merged["principal_candidates_en"] = merged["principal_candidates_en"][:4]
    merged["accessory_ingredients"] = merged["accessory_ingredients"][:8]
    return merged


def _pipeline_entries(form: dict) -> list[dict]:
    rows = []
    for e in form.get("entries") or []:
        name = str(e.get("name_ko") or e.get("name_raw") or "")
        subs = [s for s in (e.get("sub_ingredients") or []) if s]
        display = name if not subs else f"{name} ({', '.join(subs)})"
        en = str(e.get("name_en") or "").strip()
        if en:
            display = f"{display} / {en}"
        pct = e.get("percent")
        try:
            pct = float(pct) if pct not in (None, "") else None
        except (TypeError, ValueError):
            pct = None
        component = str(e.get("section") or "")
        role = _CanonicalComponentRole(
            str(e.get("role") or "other"),
            component=component,
            ingredient=display,
        )
        rows.append({
            "ingredient_name": display,
            "component": component,
            "role": role,
            "percent": pct,
            "order_index": e.get("order_index"),
            "origin": e.get("origin") or "",
            "source": "coi_normalized",
        })
    return rows


_COMPONENT_ROLE_RULES: tuple[tuple[re.Pattern[str], str], ...] = (
    (re.compile(r"(만두피|피$|도우|반죽|크러스트|wrapper|dough|crust)", re.I), "wrapper"),
    (re.compile(r"(만두소|소$|속$|필링|filling)", re.I), "filling"),
    (re.compile(r"(소스|양념|시즈닝|드레싱|비빔장|sauce|seasoning|dressing)", re.I), "sauce"),
    (re.compile(r"(육수|국물|스프|다시|broth|stock|soup\s*base)", re.I), "broth"),
    (re.compile(r"(면$|면류|생면|숙면|건면|noodle)", re.I), "noodle"),
    (re.compile(r"(고명|토핑|후레이크|건더기|topping|flake)", re.I), "topping"),
)
_INGREDIENT_LIQUID_ROLE_RE = re.compile(
    r"(소스|양념|드레싱|시즈닝|육수|국물|sauce|dressing|seasoning|broth|stock)",
    re.I,
)


def _CanonicalComponentRole(
    role: str,
    *,
    component: str,
    ingredient: str,
) -> str:
    """Backfill structural roles in legacy normalized COI forms.

    New normalized forms carry ``role`` already. Older forms frequently have
    English component labels such as ``Jjambbong Sauce`` but ``role=other``.
    Resolve only generic document vocabulary; never inspect a product name or
    assign an HS/CN code.
    """
    current = str(role or "").strip().lower()
    if current and current != "other":
        return current
    componentText = NormalizeText(component)
    for pattern, resolved in _COMPONENT_ROLE_RULES:
        if pattern.search(componentText):
            return resolved
    liquidMatch = _INGREDIENT_LIQUID_ROLE_RE.search(NormalizeText(ingredient))
    if liquidMatch:
        token = liquidMatch.group(0).lower()
        return "broth" if token in {"육수", "국물", "broth", "stock"} else "sauce"
    return "other"


def _IngredientKey(value: object) -> str:
    """Canonical key used only to apply COI-over-reconstruction precedence."""
    text = NormalizeText(value).split("/", 1)[0]
    text = re.sub(r"\([^)]*\)", "", text).lower()
    return re.sub(r"[^0-9a-z가-힣]+", "", text)


_REV_STAMP = re.compile(r"_2\d{5}(?=_|\.|$)")


def _dedupe_versions(fnames: list[str]) -> list[str]:
    """개정판 규약(설계자 07-20): 같은 서류 계보는 최신본 단일 선택.

    계보 = 파일명에서 날짜 스탬프(_2yymdd) 이후를 제거한 몸통. 같은
    계보의 복수 파일(원본+개정판)은 '추가 구성품'이 아니라 대체 관계다
    — 병합하면 성분 이중 주입(오징어무국 구판+첨가당 확인 개정판 실측).
    구성품 서류(계보 상이 — 면·소스·본품)만 병합 대상으로 남긴다.
    """
    groups: dict[str, list[str]] = {}
    for fn in fnames:
        stem = fn.rsplit(".", 1)[0]
        m = _REV_STAMP.search(stem)
        key = stem[: m.start()] if m else stem
        groups.setdefault(key, []).append(fn)

    def _rev(fn: str) -> tuple[int, str]:
        m2 = _REV_STAMP.search(fn)
        return (int(m2.group(0)[1:]) if m2 else 0, fn)

    return sorted(max(fs, key=_rev) for fs in groups.values())


def _dedupe_content(forms: list[dict]) -> list[dict]:
    """같은 서류의 재정규화 중복(스캔 변형) 제거 — 결정론.

    성분명 집합이 다른 폼의 부분집합이면 낙방(초집합 보존·동일 집합은
    선착 보존). 실측: 우동전골 '전체 구성품' 목록 2스캔 병합 → 이중.
    """
    def _names(f: dict) -> set:
        return {str(e.get("name_en") or e.get("name_ko") or "").strip().lower()
                for e in (f.get("entries") or [])} - {""}

    sets_ = [_names(f) for f in forms]
    keep = []
    for i, f in enumerate(forms):
        dominated = any(
            j != i and sets_[i] <= sets_[j]
            and (sets_[i] < sets_[j] or j < i)
            for j in range(len(forms)))
        if not dominated:
            keep.append(f)
    return keep or forms


def FindFormForProduct(product_name: str) -> dict | None:
    if not _forms_enabled():
        return None
    mapping = LoadCoiProductMap()
    fnames = mapping.get(_nfc(product_name))
    if not fnames:
        # 퍼지 폴백: 표기 차('다진 새우살' vs '다짐살 9종')로 exact 실패 시
        # 비제네릭 토큰 겹침 최다 키 (2토큰 이상일 때만 — 오매칭 방지)
        generic = {"냉동", "국산", "간편", "이유식용", "프리미엄", "오리지널"}
        ptoks = {w for w in re.findall(r"[가-힣A-Za-z]{2,}", _nfc(product_name))} - generic
        best, best_score = None, 1
        for key, fs in mapping.items():
            ktoks = {w for w in re.findall(r"[가-힣A-Za-z]{2,}", key)} - generic
            score = sum(1 for tk in ptoks if any(tk in k or k in tk for k in ktoks))
            if score > best_score:
                best, best_score = fs, score
        fnames = best
    if not fnames:
        return None
    fnames = _dedupe_versions(fnames)
    available = LoadCoiForms()
    forms = [available[fn] for fn in fnames if fn in available]
    if not forms:
        return None
    forms = _dedupe_content(forms)
    return forms[0] if len(forms) == 1 else _merge_forms(forms)


def InjectIntoProductUnderstanding(pu: dict[str, Any]) -> dict[str, Any] | None:
    """PU dict에 정규화 COI를 가산 주입. 주입 시 요약 dict, 아니면 None."""
    form = FindFormForProduct(str(pu.get("product_name") or ""))
    if not form:
        return None
    entries = _pipeline_entries(form)
    if not entries:
        return None
    cf = dict(pu.get("composition_facts") or {})
    ih = pu.get("identity_hints") or {}
    existing = [e for e in (cf.get("ingredient_entries") or []) if isinstance(e, dict)]
    # Composition authority: normalized COI first, reconstruction fills only
    # names absent from COI. Keep repeated COI ingredients across distinct
    # components because their structural role is legally meaningful.
    added: list[dict[str, Any]] = []
    coiSeen: set[tuple[str, str]] = set()
    for e in entries:
        nm = _IngredientKey(e.get("ingredient_name"))
        component = NormalizeText(e.get("component")).lower()
        key = (component, nm)
        if not nm or key in coiSeen:
            continue
        coiSeen.add(key)
        added.append(e)
    coiNames = {_IngredientKey(e.get("ingredient_name")) for e in added}
    reconstructionFillers = [
        e for e in existing
        if _IngredientKey(e.get("ingredient_name")) not in coiNames
    ]
    cf["ingredient_entries"] = added + reconstructionFillers

    # [07-20 배선] 함량 주입 — entries의 percent가 quant 소비 필드
    # (composition_facts.ingredient_percentages)로 안 옮겨지던 결손.
    # 실측: 전복미역국 재정규화로 전복 4%가 폼에 실렸는데 pct=[]라
    # 주해 2(어육 20%) 중재가 영구 불능이었다. 기존값 보존(가산).
    # [Task A] 구성품 병합 폼은 % 가 제품 레벨이 아니라 quant 미주입.
    if form.get("_constituent_merge"):
        _pct_add = []
    else:
        _pct_existing = {str(x.get("ingredient") or "").strip()
                         for x in (cf.get("ingredient_percentages") or [])
                         if isinstance(x, dict)}
        _pct_add = []
        for e in added:
            nm = str(e.get("ingredient_name") or "").strip()
            pv = e.get("percent")
            if nm and isinstance(pv, (int, float)) and nm not in _pct_existing:
                _pct_add.append({"ingredient": nm, "percent": float(pv)})
                _pct_existing.add(nm)
    if _pct_add:
        cf["ingredient_percentages"] = (
            list(cf.get("ingredient_percentages") or []) + _pct_add)

    cands = [str(c) for c in (form.get("principal_candidates") or []) if c]
    cands_en = [str(c) for c in (form.get("principal_candidates_en") or [])]
    principal = ""
    if cands:
        id_text = " ".join([
            str(ih.get("principal_ingredient_guess") or ""),
            *[str(x) for x in (ih.get("identity_terms") or [])],
            str(ih.get("normalized_tariff_description") or ""),
            str(pu.get("product_name") or ""),
        ]).lower()
        # 정체 교차가 %보다 우선 — 함량 1위(고춧가루 40%)가 본질(꼬막)을
        # 이기면 GIR 3(b)에 역행한다(꼬막장 실측). %는 교차 무성립 시의
        # 보조 신호로만, 그것도 유의미 함량(≥10%) 2개 이상 비교일 때만.
        # 한글은 복합어 표기 편차('새꼬막'/'꼬막장'/'꼬막')가 커서 토큰
        # 포함으로는 어긋난다 — 2-gram 교집합으로 판정. 영문은 토큰 포함.
        def _bigrams(text: str) -> set[str]:
            ko = re.sub(r"[^가-힣]", "", text)
            return {ko[i:i + 2] for i in range(len(ko) - 1)}

        # 구식 coi_cross_check 등가: 라벨(재구성) 1위 성분도 교차 원천에
        # 포함 — identity 서술이 흔들려도(꼬마바 'vegetable bar') 라벨
        # 1위('연육')×COI 후보('연육') 일치라는 독립 2근거가 살아남는다.
        label_first = next(
            (str(e.get("ingredient_name") or "") for e in existing
             if int(e.get("order_index") or 0) == 1),
            "",
        )
        if label_first:
            id_text = id_text + " " + label_first.lower()
        id_bi = _bigrams(id_text)
        id_en = {w for w in re.findall(r"[a-z]{3,}", id_text)}
        for ci, c in enumerate(cands):
            en = cands_en[ci] if ci < len(cands_en) else ""
            cl = (c + " " + en).lower()
            ko_hit = bool(_bigrams(cl) & id_bi)
            en_hit = any(w in id_en for w in re.findall(r"[a-z]{3,}", cl))
            if ko_hit or en_hit:
                principal = c
                break
        if not principal:
            by_pct = [(e.get("percent") or 0, str(e.get("ingredient_name") or ""))
                      for e in entries
                      if any(str(c) in str(e.get("ingredient_name") or "") for c in cands)]
            by_pct = [x for x in by_pct if x[0] and x[0] >= 10]
            if len(by_pct) >= 2:
                principal = max(by_pct)[1]
        cf["coi_principal_candidates"] = cands
    # Field-level authority: a cross-confirmed COI principal supersedes the
    # reconstruction guess. Preserve the displaced value in the audit trace.
    if principal:
        previousPrincipal = str(cf.get("principal_ingredient") or "").strip()
        cf["principal_ingredient"] = principal
        if previousPrincipal and _IngredientKey(previousPrincipal) != _IngredientKey(principal):
            traces = list(cf.get("extraction_traces") or [])
            traces.append({
                "source_field_name": "coi_normalized",
                "source_text": principal,
                "selected_span": principal,
                "output_field": "principal_ingredient",
                "normalized_value": principal,
                "extraction_method": "coi_authority_override",
                "decision_reason": (
                    "COI principal supersedes reconstruction principal "
                    f"({previousPrincipal})"
                ),
                "confidence": 1.0,
                "unresolved_reason": "",
                "source_refs": [],
            })
            cf["extraction_traces"] = traces
    if form.get("accessory_ingredients"):
        cf["accessory_ingredients"] = form["accessory_ingredients"]
    roles = {e.get("role") for e in entries}
    if "wrapper" in roles:
        cf["contains_wrapper_or_dough"] = True
    if roles & {"sauce", "broth"}:
        cf["contains_sauce_or_broth"] = True
    # 파생 필드 재계산 — 구식 로더는 lane 내부 시점이라 percentages 등이
    # COI를 반영해 계산됐지만, 본 훅은 기록 후 시점이라 재계산 없이는
    # quant 연료(ingredient_percentages)가 COI 없던 값으로 굳는다(대체런
    # 실측: 구식 OFF에서 %형 분기 연료 소실). PU의 계산기를 그대로 호출.
    try:
        from bussiness_logic.product.components.product_understanding import (
            ProductUnderstandingComponent as _PU,
        )
        # [Task A] 구성품 병합 폼은 % 재계산도 건너뛴다 — entries의
        # percent가 구성품 내부 비율이라 제품 레벨 quant로 굳으면 오독
        # (우동전골 어묵 90% → 16류 오견인). 미결 > 오판.
        if not form.get("_constituent_merge"):
            cf["ingredient_percentages"] = _PU._IngredientPercentagesFromEntries(
                cf["ingredient_entries"],
            )
        # ingredient_classes도 lane 내부 시점 계산이라 신폼 entries 미반영
        # (실측: 꼬마바 4→1, 산채 2→0 — hs2/hs4 지지 실탄 소실). 재계산.
        merged_classes = _PU._BuildIngredientClasses(
            ingredientEntries=cf["ingredient_entries"],
            compositionTerms=tuple(cf.get("composition_terms") or ()),
        )
        if merged_classes:
            cf["ingredient_classes"] = list(merged_classes)
        # 소비처 전수(2026-07-15)에서 확인된 나머지 파생 2종도 재계산 —
        # 팀장 후보 필드(dict 계약)와 status는 PU 계산기로 형식 보존.
        cand2 = _PU._BuildPrincipalCandidates(cf["ingredient_entries"])
        if cand2:
            cf["principal_ingredient_candidates"] = [dict(c) for c in cand2]
            status2 = _PU._PrincipalStatus(cand2)
            if principal:
                status2 = "confirmed"
            cf["principal_ingredient_status"] = status2
        comp2 = _PU._BuildComponentCompositions(
            productFacts=pu.get("reconstructed_product_facts") or [],
            reconstructedTables=(),
            ingredientEntries=cf["ingredient_entries"],
        )
        if comp2:
            cf["component_compositions"] = [dict(c) for c in comp2]
    except Exception:  # noqa: BLE001 — 재계산 실패는 기존 값 유지
        pass
    # [12회차-5] 하네스 주입 — 문서(COI 정규화 폼)가 실제로 답한
    # 형태·보존 상태를 direct 필드에 채운다. 값은 폼 실등장분만(창작 0),
    # provenance는 coi_normalized로 표시돼 §C 자격 심사의 입력이 된다.
    _pf = str(form.get("physical_form") or form.get("form") or "").strip()
    _ps = str(form.get("preservation_state") or form.get("storage")
              or form.get("processing_state") or "").strip()
    if _pf and not cf.get("physical_form"):
        cf["physical_form"] = _pf
    if _ps and not cf.get("preservation_state"):
        cf["preservation_state"] = _ps
    # [12회차 A] 스탬프 조건 수리 — 종전엔 형태/보존 값이 있을 때만
    # 찍었는데, COI 폼에는 그 최상위 키가 없어 신 3런 전체가 스탬프 0
    # 이었다(실측). entries 주입이 성립한 것 자체가 문서 출처이므로
    # (entry 단위 source=coi_normalized 실물과 동일 의미) 주입 성립
    # 시 항상 찍는다.
    if added or _pf or _ps:
        cf.setdefault("composition_provenance", "coi_normalized")
    pu["composition_facts"] = cf
    return {"coi_form_injected": True, "entries_added": len(added),
            "physical_form": _pf, "preservation_state": _ps,
            "principal": principal or "(후보 병기)"}
