"""COI xlsx evidence loader."""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Iterable


PROJECT_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_COI_ROOT = PROJECT_ROOT / "test" / "COI(식품원재료풀이)"
TOKEN_RE = re.compile(r"[0-9A-Za-z가-힣]+")
INDEX_PREFIX_RE = re.compile(r"^\s*([0-9,\s]+)\.")


@dataclass(frozen=True, slots=True)
class CoiEvidence:
    path: Path
    text: str
    matchedScore: int = 0


def NormalizeText(value: object) -> str:
    return unicodedata.normalize("NFC", " ".join(str(value or "").split()))


def EvidenceTokens(text: str) -> set[str]:
    return {
        token.lower()
        for token in TOKEN_RE.findall(NormalizeText(text))
        if len(token) >= 2
    }


def _IndicesFromParent(path: Path) -> set[int]:
    match = INDEX_PREFIX_RE.match(path.parent.name)
    if match is None:
        return set()
    return {
        int(chunk)
        for chunk in match.group(1).replace(" ", "").split(",")
        if chunk.isdigit()
    }


def _XlsxPaths(root: Path) -> list[Path]:
    if not root.exists():
        return []
    return sorted(
        path
        for path in root.rglob("*.xlsx")
        if not path.name.startswith("~$")
    )


def _ScorePath(path: Path, *, caseIndex: int | None, productName: str) -> int:
    score = 0
    if caseIndex is not None and caseIndex in _IndicesFromParent(path):
        score += 100
    haystack = f"{path.parent.name} {path.stem}"
    score += len(EvidenceTokens(productName) & EvidenceTokens(haystack)) * 10
    if "수출가능" in path.name:
        score += 4
    if "수정" in path.name or "최종" in path.name:
        score += 2
    if "복사" in path.name or "copy" in path.name.lower():
        score -= 1
    return score


def FindCoiPath(
    *,
    caseIndex: int | None,
    productName: str,
    coiRoot: Path = DEFAULT_COI_ROOT,
) -> Path | None:
    scored = [
        (_ScorePath(path, caseIndex=caseIndex, productName=productName), path)
        for path in _XlsxPaths(coiRoot)
    ]
    if not scored:
        return None
    scored.sort(key=lambda item: (-item[0], str(item[1])))
    score, path = scored[0]
    return path if score > 0 else None


def _CellToText(value: object) -> str:
    if value is None:
        return ""
    text = NormalizeText(value)
    if not text or text.lower() in {"none", "nan"}:
        return ""
    return text


def FlattenXlsxText(
    path: Path,
    *,
    maxChars: int = 8000,
    maxRowsPerSheet: int = 220,
) -> str:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise RuntimeError("openpyxl is required to read COI xlsx files") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet in workbook.worksheets:
            rowsAdded = 0
            parts.append(f"[sheet] {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                values = [_CellToText(value) for value in row]
                values = [value for value in values if value]
                if not values:
                    continue
                parts.append(" | ".join(values))
                rowsAdded += 1
                if rowsAdded >= maxRowsPerSheet:
                    parts.append("[sheet_truncated]")
                    break
                if sum(len(part) + 1 for part in parts) >= maxChars:
                    return "\n".join(parts)[:maxChars]
    finally:
        workbook.close()
    return "\n".join(parts)[:maxChars]


@lru_cache(maxsize=256)
def _CachedFlatten(pathText: str, maxChars: int) -> str:
    return FlattenXlsxText(Path(pathText), maxChars=maxChars)


def LoadCoiEvidence(
    *,
    caseIndex: int | None,
    productName: str,
    coiRoot: Path = DEFAULT_COI_ROOT,
    maxChars: int = 8000,
) -> CoiEvidence | None:
    path = FindCoiPath(
        caseIndex=caseIndex,
        productName=productName,
        coiRoot=coiRoot,
    )
    if path is None:
        return None
    return CoiEvidence(
        path=path,
        text=_CachedFlatten(str(path), maxChars),
        matchedScore=_ScorePath(path, caseIndex=caseIndex, productName=productName),
    )


def SummarizeCoiMatches(
    rows: Iterable[tuple[int | None, str]],
    *,
    coiRoot: Path = DEFAULT_COI_ROOT,
) -> list[dict[str, object]]:
    return [
        {
            "case_index": caseIndex,
            "product_name": productName,
            "coi_path": str(
                FindCoiPath(
                    caseIndex=caseIndex,
                    productName=productName,
                    coiRoot=coiRoot,
                )
                or ""
            ),
        }
        for caseIndex, productName in rows
    ]
